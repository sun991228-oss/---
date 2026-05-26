"""
근무성적 평정 시스템 v3
조직구조: 3부 9팀 / 직책: 대표이사·본부장·부장·팀장·팀원 / 직급: 2~6급
평가체계: 1차(팀장→팀원) / 2차(부장→소속부전체) / 3차(본부장→전직원) / 4차(대표이사→전직원)
"""

import streamlit as st
import json, hashlib, io, re
from datetime import datetime
from pathlib import Path
from collections import defaultdict
import pandas as pd

# ══════════════════════════════════════════════
# 상수 & 조직 구조
# ══════════════════════════════════════════════
ORG = {
    "경영지원부":     ["행정지원팀", "재정관리팀", "생산관리팀"],
    "직매장사업부":   ["도담점", "아름점", "새롬점", "소담점", "직매장행정팀"],
    "공공급식사업부": ["공공급식팀"],
}
ALL_DEPTS = list(ORG.keys())
ALL_TEAMS = [t for teams in ORG.values() for t in teams]
POSITIONS = ["대표이사", "본부장", "부장", "팀장", "팀원"]
GRADES    = ["2급", "3급", "4급", "5급", "6급", "공무직"]


# 직책별 평가 비중 (피평가자 직책 기준)
# 팀원: 1차(팀장30%) 2차(부장30%) 3차(본부장20%) 4차(대표이사20%)
# 팀장: 1차없음 → 2차(부장40%) 3차(본부장30%) 4차(대표이사30%)
# 부장: 1차없음 2차없음 → 3차(본부장60%) 4차(대표이사40%)
# 본부장: 1차없음 2차없음 3차없음 → 4차(대표이사100%)

EVAL_WEIGHTS_BY_POS = {
    "팀원":   {"1차": 0.3, "2차": 0.3, "3차": 0.2, "4차": 0.2},
    "공무직": {"1차": 0.3, "2차": 0.3, "3차": 0.2, "4차": 0.2},
    "팀장":   {"2차": 0.4, "3차": 0.3, "4차": 0.3},
    "부장":   {"3차": 0.6, "4차": 0.4},
    "본부장": {"4차": 1.0},
}
EVAL_WEIGHTS = {"1차": 0.3, "2차": 0.3, "3차": 0.2, "4차": 0.2}  # 기본(팀원 기준)

def get_eval_weights_for_ee(ee_pos: str) -> dict:
    """피평가자 직책에 따른 평가 비중 반환"""
    return EVAL_WEIGHTS_BY_POS.get(ee_pos, EVAL_WEIGHTS_BY_POS["팀원"])

def get_stage_order_for_ee(ee_pos: str) -> list:
    """피평가자 직책에 따른 평가 차수 순서 반환"""
    return list(get_eval_weights_for_ee(ee_pos).keys())

DEDUCTION_RATES = {
    "지각":           ("회",  0.15),
    "무단결근":       ("일",  0.50),
    "무단조퇴":       ("회",  0.25),
    "무단외출(이석)": ("회",  0.25),
    "민원야기":       ("건",  0.50),
    "경고(훈계·주의)":("건",  0.50),
    "정직":           ("건",  2.50),
    "직위해제":       ("건",  2.50),
    "감봉":           ("건",  2.00),
    "견책":           ("건",  1.50),
    "불문경고":       ("건",  1.00),
}

ABILITY_ITEMS = [
    ("직무능력", 7, ". 업무수행 단계별 일정 준수 및 보고\n. 업무수행 결과의 효과성 및 업무 완결도"),
    ("관계능력", 7, ". 회사내에서의 동료의식 및 태도\n. 유관기관, 고객, 거래처 등과 원활한 조정 및 관리"),
    ("팔로우십", 7, ". 팀장의 업무추진에 대한 지원의 적극성\n. 팀워크 활성화를 위한 기여도"),
    ("창의력",   7, ". 기존 사업에 대한 창의적 기획\n. 새로운 사업 발굴 및 만족도 제고"),
    ("소통능력", 7, ". 문서·보고서 작성을 통한 공식적인 소통 능력\n. 다른 사람의 의견을 경청하는 능력\n. 주변의 의견을 기꺼이 받아들이고 개선하는 태도"),
]

# ══════════════════════════════════════════════
# 초기화 (Supabase는 SQL로 테이블 생성 완료)
# ══════════════════════════════════════════════
# ══════════════════════════════════════════════
# Supabase 연결
# ══════════════════════════════════════════════
from supabase import create_client, Client

@st.cache_resource
def get_supabase() -> Client:
    url = st.secrets["SUPABASE_URL"]
    key = st.secrets["SUPABASE_KEY"]
    return create_client(url, key)

# ══════════════════════════════════════════════
# 유틸
# ══════════════════════════════════════════════
def hash_pw(pw): return hashlib.sha256(str(pw).encode()).hexdigest()

def is_valid_id(uid):
    return bool(uid) and bool(re.match(r'^[a-zA-Z0-9_\uAC00-\uD7A3\u3131-\u314E\u314F-\u3163]+$', str(uid)))

def get_team_dept(team):
    for dept, teams in ORG.items():
        if team in teams:
            return dept
    return ""

def grade_label(s):
    if s >= 90: return "S"
    if s >= 80: return "A"
    if s >= 70: return "B"
    if s >= 60: return "C"
    return "D"

def score_color(s):
    if s >= 80: return "#2ecc71"
    if s >= 60: return "#e67e22"
    return "#e74c3c"

# ══════════════════════════════════════════════
# DB 데이터 접근 함수 (Supabase)
# ══════════════════════════════════════════════
@st.cache_data(ttl=30)
def get_users() -> dict:
    sb = get_supabase()
    rows = sb.table("users").select("*").execute().data
    return {r["uid"]: {k: v for k, v in r.items() if k != "uid"} for r in rows}

def save_user(uid: str, data: dict):
    sb = get_supabase()
    sb.table("users").upsert({"uid": uid, **data}).execute()
    get_users.clear()

def delete_user(uid: str):
    sb = get_supabase()
    sb.table("users").delete().eq("uid", uid).execute()
    get_users.clear()

@st.cache_data(ttl=30)
def get_profiles() -> dict:
    sb = get_supabase()
    rows = sb.table("profiles").select("*").execute().data
    result = {}
    for r in rows:
        uid = r["uid"]
        result[uid] = {
            "입사일":       r.get("ipsa_date", ""),
            "현직급임용일": r.get("current_grade_date", ""),
            "현보직일":     r.get("current_pos_date", ""),
            "담당업무":     r.get("work_summary", ""),
        }
    return result

def save_profile(uid: str, p: dict):
    sb = get_supabase()
    sb.table("profiles").upsert({
        "uid":                uid,
        "ipsa_date":          p.get("입사일", ""),
        "current_grade_date": p.get("현직급임용일", ""),
        "current_pos_date":   p.get("현보직일", ""),
        "work_summary":       p.get("담당업무", ""),
        "updated_at":         datetime.now().isoformat(),
    }).execute()
    get_profiles.clear()

@st.cache_data(ttl=30)
def get_tasks() -> dict:
    sb = get_supabase()
    rows = sb.table("tasks").select("*").execute().data
    return {r["uid"]: {"tasks": r.get("tasks_json", []),
                        "updated": r.get("updated_at", "")} for r in rows}

def save_tasks(uid: str, tasks_list: list):
    sb = get_supabase()
    sb.table("tasks").upsert({
        "uid":        uid,
        "tasks_json": tasks_list,
        "updated_at": datetime.now().isoformat(),
    }).execute()
    get_tasks.clear()

@st.cache_data(ttl=30)
def get_evaluations() -> dict:
    sb = get_supabase()
    ev_rows = sb.table("evaluations").select("*").execute().data
    gr_rows = sb.table("assigned_grades").select("*").execute().data

    result: dict = {}
    for r in ev_rows:
        uid   = r["uid"]
        stage = r["stage"]
        result.setdefault(uid, {})[stage] = r.get("data_json", {})

    for r in gr_rows:
        uid = r["uid"]
        result.setdefault(uid, {})["assigned_grade"]    = r.get("grade", "미확정")
        result[uid]["grade_assigned_by"] = r.get("assigned_by", "")
        result[uid]["grade_assigned_at"] = str(r.get("assigned_at", ""))

    return result

def save_evaluation(uid: str, stage: str, data: dict):
    sb = get_supabase()
    sb.table("evaluations").upsert({
        "uid":        uid,
        "stage":      stage,
        "data_json":  data,
        "updated_at": datetime.now().isoformat(),
    }).execute()
    get_evaluations.clear()

def save_assigned_grade(uid: str, grade: str, assigned_by: str):
    sb = get_supabase()
    sb.table("assigned_grades").upsert({
        "uid":         uid,
        "grade":       grade,
        "assigned_by": assigned_by,
        "assigned_at": datetime.now().isoformat(),
    }).execute()
    get_evaluations.clear()

def get_evaluatees() -> dict:
    return {k: v for k, v in get_users().items() if v.get("role") == "evaluatee"}

def get_evaluators() -> dict:
    return {k: v for k, v in get_users().items() if v.get("role") == "evaluator"}

# PDF는 Supabase Storage 사용
BUCKET = "eval-pdfs"

def upload_pdf(uid: str, file_bytes: bytes):
    sb = get_supabase()
    path = f"{uid}.pdf"
    try:
        sb.storage.from_(BUCKET).remove([path])
    except Exception:
        pass
    sb.storage.from_(BUCKET).upload(path, file_bytes, {"content-type": "application/pdf"})

def get_pdf_url(uid: str) -> str | None:
    sb = get_supabase()
    try:
        files = sb.storage.from_(BUCKET).list()
        names = [f["name"] for f in files]
        if f"{uid}.pdf" in names:
            return sb.storage.from_(BUCKET).get_public_url(f"{uid}.pdf")
    except Exception:
        pass
    return None

@st.cache_data(ttl=60)
def get_pdf_names() -> set:
    """Storage에 있는 PDF 파일명 목록을 한 번에 가져옴"""
    sb = get_supabase()
    try:
        files = sb.storage.from_(BUCKET).list()
        return {f["name"] for f in files}
    except Exception:
        return set()

def pdf_exists(uid: str) -> bool:
    return f"{uid}.pdf" in get_pdf_names()

def delete_pdf(uid: str):
    sb = get_supabase()
    try:
        sb.storage.from_(BUCKET).remove([f"{uid}.pdf"])
    except Exception:
        pass
    get_pdf_names.clear()

def reset_evaluation(uid: str):
    """피평가자의 모든 평가 데이터 초기화 (과제·인적사항·평가·등급)"""
    sb = get_supabase()
    try:
        sb.table("evaluations").delete().eq("uid", uid).execute()
    except Exception:
        pass
    try:
        sb.table("assigned_grades").delete().eq("uid", uid).execute()
    except Exception:
        pass
    try:
        sb.table("tasks").delete().eq("uid", uid).execute()
    except Exception:
        pass
    try:
        sb.table("profiles").delete().eq("uid", uid).execute()
    except Exception:
        pass
    delete_pdf(uid)
    get_users.clear()
    get_evaluations.clear()
    get_tasks.clear()
    get_profiles.clear()

def get_selfreport(uid: str) -> dict:
    sb = get_supabase()
    try:
        rows = sb.table("selfreports").select("*").eq("uid", uid).execute().data
        if rows:
            r = rows[0]
            return {
                "dev1":       r.get("dev1", ""),
                "dev2":       r.get("dev2", ""),
                "goals":      r.get("goals_json", ["","","","",""]),
                "suggestion": r.get("suggestion", ""),
                "updated_at": r.get("updated_at", ""),
            }
    except Exception:
        pass
    return {"dev1":"","dev2":"","goals":["","","","",""],"suggestion":"","updated_at":""}

def save_selfreport(uid: str, data: dict):
    sb = get_supabase()
    sb.table("selfreports").upsert({
        "uid":         uid,
        "dev1":        data.get("dev1",""),
        "dev2":        data.get("dev2",""),
        "goals_json":  data.get("goals",[]),
        "suggestion":  data.get("suggestion",""),
        "updated_at":  data.get("updated_at", datetime.now().isoformat()),
    }).execute()


# ══════════════════════════════════════════════
# 조직 기반 평가 범위
# ══════════════════════════════════════════════
def get_stage_for_evaluator(uid):
    pos = get_users().get(uid, {}).get("position", "")
    return {"팀장": "1차", "부장": "2차", "본부장": "3차", "대표이사": "4차"}.get(pos, "")

def get_evaluatee_scope(ev_uid):
    """평가자가 담당하는 피평가자 목록 반환 (직책별 평가 범위)"""
    users    = get_users()
    profiles = get_profiles()
    ev       = users.get(ev_uid, {})
    ev_pos   = ev.get("position", "")
    ev_dept  = ev.get("dept", "")
    ev_team  = ev.get("team", "")
    result   = []

    for uid, u in users.items():
        if u.get("role") != "evaluatee":
            continue
        p       = profiles.get(uid, {})
        ee_team = p.get("team", u.get("team", ""))
        ee_dept = get_team_dept(ee_team)
        ee_pos  = p.get("position", u.get("position", "팀원"))

        if ev_pos == "팀장":
            # 자기 팀 팀원만 (팀장 본인의 1차 평가 대상)
            if ee_team == ev_team and ee_pos in ("팀원", "공무직"):
                result.append(uid)
        elif ev_pos == "부장":
            # 소속 부 전체: 팀원+팀장 (부장은 2차 평가자)
            if ee_dept == ev_dept and ee_pos in ("팀원", "팀장", "공무직"):
                result.append(uid)
        elif ev_pos == "본부장":
            # 전사: 팀원+팀장+부장 (본부장은 3차 평가자)
            if ee_pos in ("팀원", "팀장", "부장", "공무직"):
                result.append(uid)
        elif ev_pos == "대표이사":
            # 전사 모든 피평가자 (4차 평가자)
            result.append(uid)

    return result

# ══════════════════════════════════════════════
# 점수 계산
# ══════════════════════════════════════════════
def calc_task_score(tasks, ev_stage_data):
    return round(sum(ev_stage_data.get("tasks", {}).get(t["id"], 0) * t["weight"] * 6
                     for t in tasks), 2)

def calc_ability_score(ev_stage_data):
    ab = ev_stage_data.get("ability", {})
    return round(sum(ab.get(n, 0) for n, _, _ in ABILITY_ITEMS), 2)

def calc_attitude_score(deductions):
    total = sum(deductions.get(item, 0) * rate for item, (_, rate) in DEDUCTION_RATES.items())
    return round(max(0.0, 5.0 - total), 2)

def calc_final(ee_id, evaluations, tasks_data):
    tasks      = tasks_data.get(ee_id, {}).get("tasks", [])
    deductions = evaluations.get(ee_id, {}).get("deductions", {})
    att_c      = calc_attitude_score(deductions)
    result     = {}
    w_a = w_b = w_ab = 0.0
    all_done   = True

    for stage, weight in EVAL_WEIGHTS.items():
        ev_data = evaluations.get(ee_id, {}).get(stage, {})
        if not ev_data:
            all_done = False
            continue
        a = calc_task_score(tasks, ev_data)
        b = calc_ability_score(ev_data)
        result[stage] = {"A": a, "B": b, "AB": round(a+b,2), "weight": weight}
        w_a  += a * weight
        w_b  += b * weight
        w_ab += (a+b) * weight

    ded_done = evaluations.get(ee_id, {}).get("deductions") is not None
    result["종합"] = {
        "A_가중":  round(w_a,2),  "B_가중": round(w_b,2),
        "AB_가중": round(w_ab,2), "C": att_c,
        "최종":    round(w_ab + att_c, 2),
        "완료":    all_done and ded_done,
    }
    return result

def calc_grade_rankings(evaluations, tasks_data):
    users    = get_users()
    profiles = get_profiles()
    rows = []
    for uid, u in users.items():
        if u.get("role") != "evaluatee":
            continue
        p      = profiles.get(uid, {})
        ee_pos = p.get("position", u.get("position", "팀원"))
        # 직책 기준 평가 차수 목록
        ee_stages = get_stage_order_for_ee(ee_pos)
        ev     = evaluations.get(uid, {})
        tasks  = tasks_data.get(uid, {}).get("tasks", [])

        # 해당 직책의 모든 차수가 완료됐는지 확인
        if not all(ev.get(s) for s in ee_stages):
            continue
        # 직무수행태도 완료 여부 (팀원·공무직만 필수)
        if ee_pos in ("팀원", "공무직") and ev.get("deductions") is None:
            continue

        result = calc_final(uid, evaluations, tasks_data)
        종합   = result.get("종합", {})

        rows.append({
            "uid":      uid,
            "이름":     u.get("name",""),
            "소속부":   get_team_dept(p.get("team", u.get("team",""))),
            "소속팀":   p.get("team", u.get("team","")),
            "직책":     ee_pos,
            "직급":     p.get("grade", u.get("grade","")),
            "최종점수": 종합.get("최종", 0),
            **{f"{st_} 환산": result.get(st_, {}).get("AB", "-") for st_ in ee_stages},
        })
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows)
    df["직급순위"] = df.groupby("직급")["최종점수"].rank(method="min", ascending=False).astype(int)
    return df.sort_values(["직급", "직급순위"]).reset_index(drop=True)

# ══════════════════════════════════════════════
# 초기화
# ══════════════════════════════════════════════


for _k, _v in [("logged_in",False),("username",""),("role",""),("name","")]:
    if _k not in st.session_state: st.session_state[_k] = _v

# ══════════════════════════════════════════════════════════════
# 로그인
# ══════════════════════════════════════════════════════════════
def show_login():
    st.markdown("""
    <div style='text-align:center;padding:2.5rem 0 1rem'>
      <h1 style='font-size:2.2rem;color:#1F4E79'>📋 근무성적 평정 시스템</h1>
      <p style='color:#888'>계정 유형에 맞게 로그인하세요</p>
    </div>""", unsafe_allow_html=True)
    _, col, _ = st.columns([1,2,1])
    with col:
        with st.container(border=True):
            st.subheader("🔐 로그인")
            uid = st.text_input("아이디")
            pw  = st.text_input("비밀번호", type="password")
            if st.button("로그인", type="primary", use_container_width=True):
                users = get_users()
                input_hash = hash_pw(pw)
                db_hash = users.get(uid, {}).get("password", "")
                if uid in users and db_hash == input_hash:
                    for k,v in [("logged_in",True),("username",uid),
                                 ("role",users[uid]["role"]),("name",users[uid]["name"])]:
                        st.session_state[k] = v
                    st.rerun()
                else:
                    st.error("아이디 또는 비밀번호가 올바르지 않습니다.")

# ══════════════════════════════════════════════════════════════
# 피평가자
# ══════════════════════════════════════════════════════════════
def show_evaluatee():
    uid      = st.session_state.username
    name     = st.session_state.name
    u        = get_users().get(uid, {})
    profiles = get_profiles()
    p        = profiles.get(uid, {})
    team     = p.get("team", u.get("team",""))
    dept     = get_team_dept(team)
    pos      = p.get("position", u.get("position",""))
    grd      = p.get("grade", u.get("grade",""))

    st.title("📄 내 근무성적 평정")
    st.caption(f"👤 {name} | {dept} {team} | {pos} ({grd})")
    st.divider()

    t1, t2, t3 = st.tabs(["👤 인적사항","📝 담당업무·과제","📋 평가 참고자료 입력"])

    with t1:
        st.subheader("인적 사항")
        st.info(f"소속: **{dept} / {team}** | 직책: **{pos}** | 직급: **{grd}**\n\n*(조직 정보는 총괄 관리자가 관리합니다)*")
        with st.form("profile_form"):
            c1, c2 = st.columns(2)
            ipsa     = c1.text_input("입사일",       value=p.get("입사일",""),       placeholder="예: 2020-03-02")
            hyun_jik = c1.text_input("현직급임용일", value=p.get("현직급임용일",""), placeholder="예: 2023-01-01")
            hyun_bo  = c1.text_input("현보직일",     value=p.get("현보직일",""),     placeholder="예: 2024-03-01")
            damdan   = c2.text_input("담당업무(요약)",value=p.get("담당업무",""))
            if st.form_submit_button("💾 저장", type="primary"):
                save_profile(uid, {**p, "입사일":ipsa,"현직급임용일":hyun_jik,"현보직일":hyun_bo,"담당업무":damdan})
                st.success("저장되었습니다.")

    with t2:
        st.subheader("담당업무·과제 등록")
        st.caption("개별과제(최대 5개, 비중 합계 90%) + 팀별과제(1개, 비중 10%) / 업무비중 합계 = 100%")
        tasks_data = get_tasks()
        my_tasks   = tasks_data.get(uid, {}).get("tasks", [])
        with st.form("task_form"):
            st.markdown("**개별과제**")
            indiv_prev = [t for t in my_tasks if t.get("type")=="개별"]
            team_prev  = next((t for t in my_tasks if t.get("type")=="팀별"), {})
            indiv_new  = []
            for i in range(5):
                pr = indiv_prev[i] if i < len(indiv_prev) else {}
                c1, c2, c3 = st.columns([3,1,4])
                title      = c1.text_input(f"과제명 {i+1}", value=pr.get("title",""), key=f"it{i}")
                weight_pct = c2.number_input("비중(%)", 0, 90,
                                             int(round(float(pr.get("weight",0.0)) * 100)),
                                             5, format="%d", key=f"iw{i}")
                res        = c3.text_area("주요실적", value=pr.get("result",""), height=60, key=f"ir{i}")
                if title:
                    indiv_new.append({"id":f"indiv_{i}","type":"개별","no":i+1,
                                      "title":title,"weight":weight_pct/100,"result":res})
            st.markdown("**팀별과제**")
            c1, c2, c3 = st.columns([3,1,4])
            t_title  = c1.text_input("팀별과제명", value=team_prev.get("title",""))
            c2.number_input("비중(%)", 10, 10, 10, 5, format="%d", key="tw", disabled=True)
            t_res    = c3.text_area("주요실적", value=team_prev.get("result",""), height=60)
            team_new = [{"id":"team_0","type":"팀별","no":1,
                         "title":t_title,"weight":0.1,"result":t_res}] if t_title else []
            all_tasks = indiv_new + team_new
            indiv_w_pct = round(sum(t["weight"] for t in indiv_new) * 100)
            ok = indiv_w_pct == 90 and len(team_new) == 1
            c1, c2 = st.columns(2)
            c1.metric("개별과제 비중 합계", f"{indiv_w_pct}% / 90%", delta="✅" if indiv_w_pct==90 else "❌")
            c2.metric("팀별과제 비중", "10% (고정)", delta="✅" if team_new else "팀별과제 미입력")
            if st.form_submit_button("💾 과제 저장", type="primary"):
                if indiv_w_pct != 90:
                    st.error(f"개별과제 비중 합계가 90%이어야 합니다. (현재 {indiv_w_pct}%)")
                elif not team_new:
                    st.error("팀별과제를 반드시 1개 입력하세요.")
                elif not indiv_new:
                    st.error("과제를 최소 1개 입력하세요.")
                else:
                    save_tasks(uid, all_tasks)
                    st.success("저장되었습니다."); st.rerun()
        if my_tasks:
            st.divider()
            hrow = st.columns([1, 1, 3, 1, 5])
            hrow[0].markdown("**구분**"); hrow[1].markdown("**번호**")
            hrow[2].markdown("**과제명**"); hrow[3].markdown("**비중**")
            hrow[4].markdown("**주요실적**")
            st.divider()
            for i, t in enumerate(my_tasks):
                trow = st.columns([1, 1, 3, 1, 5])
                trow[0].caption("🔹개별" if t["type"]=="개별" else "🔸팀별")
                trow[1].caption(str(t["no"]))
                trow[2].markdown(t["title"])
                trow[3].caption(f"{t['weight']:.0%}")
                result_html = t.get("result","").replace("\n","<br>")
                trow[4].markdown(
                    f"<div style='font-size:0.85rem;line-height:1.5'>{result_html}</div>",
                    unsafe_allow_html=True
                )
                if i < len(my_tasks) - 1:
                    st.divider()

    # ── 평가 참고자료 입력 ────────────────────
    with t3:
        st.subheader("📋 평가 참고자료 입력")
        st.caption("평가자가 참고자료로 열람합니다. 작성 후 저장하세요.")

        selfreport = get_selfreport(uid)

        with st.form(f"selfreport_form_{uid}"):
            st.markdown("#### □ 자기계발 사항")
            dev1 = st.text_area(
                "1. 자기계발 및 경력관리를 위해 교육연수를 받거나 연구한 사항",
                value=selfreport.get("dev1",""),
                height=120,
                placeholder="올해 이수한 교육, 자격증 취득, 연구 내용 등을 기술하세요.",
            )
            dev2 = st.text_area(
                "2. 차기년도 자기계발 및 경력관리를 위해 교육연수를 받거나 연구하고 싶은 사항",
                value=selfreport.get("dev2",""),
                height=120,
                placeholder="내년도 희망 교육연수, 취득하고 싶은 자격증 등을 기술하세요.",
            )

            st.divider()
            st.markdown("#### □ 다음연도 고과평가 대상 기간 중 추진하고자 하는 업무목표")
            st.caption("5개 이내의 주요 업무목표를 설정하세요.")
            goals = []
            prev_goals = selfreport.get("goals", ["","","","",""])
            if len(prev_goals) < 5:
                prev_goals += [""] * (5 - len(prev_goals))
            for i in range(5):
                g = st.text_input(f"업무목표 {i+1}", value=prev_goals[i],
                                  key=f"sr_goal_{uid}_{i}",
                                  placeholder=f"업무목표 {i+1}을 입력하세요.")
                goals.append(g)

            st.divider()
            st.markdown("#### □ 희망부서 및 건의사항")
            st.caption("기타 인사상의 건의, 요망사항(희망부서 등)을 기술하세요.")
            suggestion = st.text_area(
                "희망부서 및 건의사항",
                value=selfreport.get("suggestion",""),
                height=120,
                placeholder="희망 부서, 인사 관련 건의사항 등을 자유롭게 기술하세요.",
                label_visibility="collapsed",
            )

            if st.form_submit_button("💾 저장", type="primary", use_container_width=True):
                save_selfreport(uid, {
                    "dev1": dev1, "dev2": dev2,
                    "goals": goals, "suggestion": suggestion,
                    "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
                })
                st.success("✅ 평가 참고자료가 저장되었습니다.")
                st.rerun()

        # 저장일시 표시
        if selfreport.get("updated_at"):
            st.caption(f"마지막 저장: {selfreport.get('updated_at','')}")


# ══════════════════════════════════════════════════════════════
# 평가자
# ══════════════════════════════════════════════════════════════
def show_evaluator():
    ev_uid  = st.session_state.username
    ev_name = st.session_state.name
    users   = get_users()
    ev_pos  = users.get(ev_uid, {}).get("position", "")
    stage   = get_stage_for_evaluator(ev_uid)

    st.title(f"✍️ {stage} 평가")
    st.caption(f"👤 {ev_name} | 직책: {ev_pos} | 담당 차수: {stage}")
    st.divider()

    tasks_data  = get_tasks()
    evaluations = get_evaluations()
    profiles    = get_profiles()
    assigned    = get_evaluatee_scope(ev_uid)

    if not assigned:
        st.warning("담당 피평가자가 없습니다. 총괄 관리자에게 문의하세요."); return

    stage_order = list(EVAL_WEIGHTS.keys())  # 전체 차수 순서 (1차~4차)

    # 팀별 그룹핑
    group = defaultdict(list)
    for ee_id in assigned:
        ee_u = users.get(ee_id, {})
        ee_p = profiles.get(ee_id, {})
        team = ee_p.get("team", ee_u.get("team","기타"))
        group[team].append(ee_id)

    tab_names = list(group.keys())
    tabs = st.tabs(tab_names) if tab_names else []

    for tab, team_name in zip(tabs, tab_names):
        with tab:
            for ee_id in group[team_name]:
                if ee_id not in users: continue
                ee_u    = users[ee_id]
                ee_p    = profiles.get(ee_id, {})
                ee_name = ee_u.get("name", ee_id)
                ee_pos  = ee_p.get("position", ee_u.get("position","팀원"))
                ee_grd  = ee_p.get("grade", ee_u.get("grade",""))
                tasks   = tasks_data.get(ee_id, {}).get("tasks", [])
                ev_all  = evaluations.get(ee_id, {})
                my_ev   = ev_all.get(stage, {})

                # 피평가자 직책에 맞는 평가 차수 목록
                ee_stages = get_stage_order_for_ee(ee_pos)
                # 현재 차수가 이 피평가자의 평가 대상인지 확인
                if stage not in ee_stages:
                    continue  # 이 피평가자는 현재 차수 평가 대상 아님

                # 이전 차수 잠금 (피평가자 직책 기준)
                stage_idx  = ee_stages.index(stage)
                prev_stage = ee_stages[stage_idx - 1] if stage_idx > 0 else None

                with st.expander(f"👤 {ee_name}  ({ee_pos} / {ee_grd})", expanded=False):
                    if prev_stage and not ev_all.get(prev_stage):
                        st.warning(f"⏳ {prev_stage} 평가가 완료되지 않았습니다."); continue

                    # ── 이전 차수 평가 결과 확인 ──────────────────
                    prev_stages_done = [s for s in ee_stages
                                        if s != stage and
                                        ee_stages.index(s) < ee_stages.index(stage)
                                        and ev_all.get(s)]
                    if prev_stages_done:
                        with st.expander(f"📋 이전 평가 결과 보기 ({', '.join(prev_stages_done)})", expanded=False):
                            p_cols = st.columns(len(prev_stages_done))
                            for col, ps in zip(p_cols, prev_stages_done):
                                ps_data = ev_all.get(ps, {})
                                ev_user = users.get(ps_data.get("evaluator_id",""), {})
                                a = calc_task_score(tasks, ps_data)
                                b = calc_ability_score(ps_data)
                                with col:
                                    st.markdown(f"**{ps} 평가결과**")
                                    st.caption(f"평가자: {ev_user.get('name', ps_data.get('evaluator_id',''))}")
                                    st.caption(f"평가일: {ps_data.get('date','')}")
                                    st.markdown(
                                        f"<div style='text-align:center;padding:.5rem;border-radius:8px;"
                                        f"background:#1F4E7915;border:1.5px solid #1F4E79;margin:.2rem 0'>"
                                        f"<div style='font-size:.72rem;color:#555'>근무실적(A)</div>"
                                        f"<div style='font-size:1.3rem;font-weight:bold;color:#1F4E79'>{a:.1f}점</div>"
                                        f"</div>"
                                        f"<div style='text-align:center;padding:.5rem;border-radius:8px;"
                                        f"background:#37562315;border:1.5px solid #375623;margin:.2rem 0'>"
                                        f"<div style='font-size:.72rem;color:#555'>직무능력(B)</div>"
                                        f"<div style='font-size:1.3rem;font-weight:bold;color:#375623'>{b:.1f}점</div>"
                                        f"</div>"
                                        f"<div style='text-align:center;padding:.5rem;border-radius:8px;"
                                        f"background:#C0000015;border:1.5px solid #C00000;margin:.2rem 0'>"
                                        f"<div style='font-size:.72rem;color:#555'>소계(A+B)</div>"
                                        f"<div style='font-size:1.3rem;font-weight:bold;color:#C00000'>{round(a+b,1):.1f}점</div>"
                                        f"</div>", unsafe_allow_html=True)
                                    if tasks:
                                        st.markdown("**과제별 점수**")
                                        for t in tasks:
                                            sc   = ps_data.get("tasks",{}).get(t["id"],"-")
                                            conv = round(sc * t["weight"] * 6, 2) if isinstance(sc,(int,float)) else "-"
                                            st.caption(f"{'🔹' if t['type']=='개별' else '🔸'} {t['title']}: {sc}점 → {conv}점")
                                    st.markdown("**직무수행능력**")
                                    for ab_name, ab_max, _ in ABILITY_ITEMS:
                                        ab_sc = ps_data.get("ability",{}).get(ab_name,"-")
                                        st.caption(f"{ab_name}: {ab_sc}/{ab_max}점")
                                    op = ps_data.get("opinion","")
                                    if op:
                                        st.markdown("**평정 의견**")
                                        st.info(op)
                        st.divider()

                    col_info = st.container()
                    with col_info:
                        if tasks:
                            # 헤더
                            hrow = st.columns([1, 3, 1, 5])
                            hrow[0].markdown("**구분**")
                            hrow[1].markdown("**과제명**")
                            hrow[2].markdown("**비중**")
                            hrow[3].markdown("**주요실적**")
                            st.divider()
                            for i, t in enumerate(tasks):
                                trow = st.columns([1, 3, 1, 5])
                                trow[0].caption("🔹개별" if t["type"]=="개별" else "🔸팀별")
                                trow[1].markdown(t["title"])
                                trow[2].caption(f"{t['weight']:.0%}")
                                # 줄바꿈 보존: \n → <br>
                                result_html = t.get("result","").replace("\n","<br>")
                                trow[3].markdown(
                                    f"<div style='font-size:0.85rem;line-height:1.5'>{result_html}</div>",
                                    unsafe_allow_html=True
                                )
                                if i < len(tasks) - 1:
                                    st.divider()
                        else:
                            st.caption("등록된 과제가 없습니다.")

                    # 근무성적평정서 참고자료
                    sr = get_selfreport(ee_id)
                    if any([sr.get("dev1"), sr.get("dev2"),
                            any(sr.get("goals",[])), sr.get("suggestion")]):
                        with st.expander("📋 평가 참고자료 보기", expanded=False):
                            if sr.get("dev1"):
                                st.markdown("**자기계발 1. 올해 교육연수**")
                                st.info(sr["dev1"])
                            if sr.get("dev2"):
                                st.markdown("**자기계발 2. 내년도 희망 교육연수**")
                                st.info(sr["dev2"])
                            goals = [g for g in sr.get("goals",[]) if g]
                            if goals:
                                st.markdown("**다음연도 업무목표**")
                                for i, g in enumerate(goals, 1):
                                    st.markdown(f"{i}. {g}")
                            if sr.get("suggestion"):
                                st.markdown("**희망부서 및 건의사항**")
                                st.info(sr["suggestion"])
                            st.caption(f"작성일: {sr.get('updated_at','')}")
                    else:
                        st.caption("📋 평가 참고자료 미작성")
                    with st.form(f"eval_{stage}_{ee_id}"):

                        # ① 근무실적 평정 (A, 60점)
                        st.markdown("#### ① 근무실적 평정 (A, 60점)")
                        st.caption("각 과제별 10점 만점으로 평가 → 환산점수 = 평가점수 × 업무비중(%) × 6")

                        task_scores = {}
                        task_total_conv = 0.0

                        if not tasks:
                            st.warning("피평가자가 과제를 등록하지 않았습니다.")
                        else:
                            hcols = st.columns([3, 1, 1, 1])
                            hcols[0].markdown("**과제명**")
                            hcols[1].markdown("**비중**")
                            hcols[2].markdown("**평가(10점)**")
                            hcols[3].markdown("**환산점수**")
                            for t in tasks:
                                tid    = t["id"]
                                prev_s = my_ev.get("tasks", {}).get(tid, 0)
                                rc = st.columns([3, 1, 1, 1])
                                rc[0].markdown(f"{'🔹' if t['type']=='개별' else '🔸'} {t['title']}")
                                rc[1].markdown(f"**{t['weight']:.0%}**")
                                sc = rc[2].number_input(
                                    "", min_value=0, max_value=10,
                                    value=int(prev_s), step=1,
                                    key=f"ts_{stage}_{ee_id}_{tid}",
                                    label_visibility="collapsed"
                                )
                                conv = round(sc * t["weight"] * 6, 2)
                                rc[3].markdown(f"**{conv:.2f}점**")
                                task_scores[tid] = sc
                                task_total_conv += conv
                            st.info(f"➡️ 근무실적 합계: **{task_total_conv:.2f}점** / 60점")

                        st.divider()

                        # ② 직무수행능력 (B, 35점)
                        st.markdown("#### ② 직무수행능력 (B, 35점)")
                        st.caption("각 항목별 배점 만점으로 평가하세요.")

                        ability_scores = {}
                        ab_total = 0

                        hcols2 = st.columns([2, 1, 4, 1])
                        hcols2[0].markdown("**평정요소**")
                        hcols2[1].markdown("**배점**")
                        hcols2[2].markdown("**정의**")
                        hcols2[3].markdown("**점수**")

                        for ab_name, ab_max, ab_def in ABILITY_ITEMS:
                            prev_ab = my_ev.get("ability", {}).get(ab_name, 0)
                            row2 = st.columns([2, 1, 4, 1])
                            row2[0].markdown(f"**{ab_name}**")
                            row2[1].markdown(f"{ab_max}점")
                            row2[2].markdown(
                                f"<small>{ab_def.replace(chr(10), '<br>')}</small>",
                                unsafe_allow_html=True
                            )
                            ab_sc = row2[3].number_input(
                                "", min_value=0, max_value=ab_max,
                                value=int(prev_ab), step=1,
                                key=f"ab_{stage}_{ee_id}_{ab_name}",
                                label_visibility="collapsed"
                            )
                            ability_scores[ab_name] = ab_sc
                            ab_total += ab_sc

                        st.info(f"➡️ 직무수행능력 합계: **{ab_total}점** / 35점")

                        st.divider()
                        prev_op = my_ev.get("opinion","")
                        opinion = st.text_area("📝 평정 의견", value=prev_op, height=80,
                                               placeholder="종합 의견을 작성하세요.",
                                               key=f"op_{stage}_{ee_id}")
                        submitted = st.form_submit_button("💾 평가 저장", type="primary",
                                                           use_container_width=True)

                    if submitted:
                        save_evaluation(ee_id, stage, {
                            "tasks":        task_scores,
                            "ability":      ability_scores,
                            "opinion":      opinion,
                            "evaluator_id": ev_uid,
                            "date":         datetime.now().strftime("%Y-%m-%d %H:%M"),
                        })
                        st.success(f"✅ {ee_name}의 {stage} 평가 저장 완료!"); st.rerun()

                    if my_ev:
                        st.caption(f"마지막 저장: {my_ev.get('date','')} | {my_ev.get('evaluator_id','')}")

# ══════════════════════════════════════════════════════════════
# 총괄 관리자
# ══════════════════════════════════════════════════════════════
def show_admin():
    st.title("🗂️ 평가 총괄 관리")
    st.caption(f"👤 {st.session_state.name} | 총괄 관리자")
    st.divider()

    evaluations = get_evaluations()
    tasks_data  = get_tasks()
    users       = get_users()
    profiles    = get_profiles()
    evaluatees  = get_evaluatees()

    total   = len(evaluatees)
    pdf_ok  = sum(1 for u in evaluatees if pdf_exists(u))
    task_ok = sum(1 for u in evaluatees if tasks_data.get(u,{}).get("tasks"))
    done_4  = sum(1 for u in evaluatees if evaluations.get(u,{}).get("4차"))
    ded_ok  = sum(1 for u in evaluatees if evaluations.get(u,{}).get("deductions") is not None)
    full_ok = sum(1 for u in evaluatees
                  if calc_final(u, evaluations, tasks_data).get("종합",{}).get("완료"))

    cols = st.columns(6)
    for c, (label, val) in zip(cols, [
        ("전체 피평가자",f"{total}명"),("과제 등록",f"{task_ok}명"),
        ("PDF 제출",f"{pdf_ok}명"),("4차 완료",f"{done_4}명"),
        ("태도 입력",f"{ded_ok}명"),("전체 완료",f"{full_ok}명")
    ]):
        c.metric(label, val)

    st.divider()

    tab1, tab2, tab3, tab4 = st.tabs([
        "📋 전체 현황", "🧘 직무수행태도", "🏆 직급별 순위", "🧑‍💼 계정 관리"
    ])

    # ── 탭1 전체 현황 ─────────────────────────
    with tab1:
        st.subheader("피평가자별 평가 진행 현황")
        col_fd, col_ft = st.columns(2)
        filter_dept = col_fd.selectbox("부 필터", ["전체"]+ALL_DEPTS, key="fd")
        filter_team = col_ft.selectbox("팀 필터", ["전체"]+ALL_TEAMS, key="ft")

        filtered = []
        for uid, u in evaluatees.items():
            p    = profiles.get(uid, {})
            team = p.get("team", u.get("team",""))
            dept = get_team_dept(team)
            if filter_dept != "전체" and dept != filter_dept: continue
            if filter_team != "전체" and team != filter_team: continue
            filtered.append((uid, u, p, team, dept))

        if not filtered:
            st.info("해당 조건의 피평가자가 없습니다.")
        else:
            # 테이블 헤더
            hcols = st.columns([2, 1.5, 1.5, 1, 1, 1, 1, 1, 1, 1, 1.2, 1.2])
            for hc, label in zip(hcols, ["이름","소속부","소속팀","직책","직급",
                                          "과제","1차","2차","3차","4차","태도","평정표"]):
                hc.markdown(f"**{label}**")
            st.divider()

            rows_csv = []
            for uid, u, p, team, dept in filtered:
                ev   = evaluations.get(uid, {})
                ag   = ev.get("assigned_grade", "-")
                rc   = st.columns([2, 1.5, 1.5, 1, 1, 1, 1, 1, 1, 1, 1.2, 1.2])
                rc[0].markdown(f"**{u.get('name','')}**")
                rc[1].caption(dept)
                rc[2].caption(team)
                rc[3].caption(p.get("position",u.get("position","")))
                rc[4].caption(p.get("grade",u.get("grade","")))
                rc[5].caption("✅" if tasks_data.get(uid,{}).get("tasks") else "❌")
                rc[6].caption("✅" if ev.get("1차") else "⏳")
                rc[7].caption("✅" if ev.get("2차") else "⏳")
                rc[8].caption("✅" if ev.get("3차") else "⏳")
                rc[9].caption("✅" if ev.get("4차") else "⏳")
                rc[10].caption("✅" if ev.get("deductions") is not None else "⏳")
                xl_bytes = generate_eval_excel(uid, u, evaluations, tasks_data)
                rc[11].download_button("⬇️", data=xl_bytes,
                                       file_name=f"{u.get('name','')}_근무성적평정표.xlsx",
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                       key=f"dl_{uid}")
                rows_csv.append({
                    "이름": u.get("name",""), "소속부": dept, "소속팀": team,
                    "직책": p.get("position",u.get("position","")),
                    "직급": p.get("grade",u.get("grade","")),
                    "과제등록": "✅" if tasks_data.get(uid,{}).get("tasks") else "❌",
                    "1차": "✅" if ev.get("1차") else "⏳",
                    "2차": "✅" if ev.get("2차") else "⏳",
                    "3차": "✅" if ev.get("3차") else "⏳",
                    "4차": "✅" if ev.get("4차") else "⏳",
                    "태도입력": "✅" if ev.get("deductions") is not None else "⏳",
                    "확정등급": ag,
                })

            st.divider()
            df_csv = pd.DataFrame(rows_csv)
            col_dl1, col_dl2 = st.columns(2)
            col_dl1.download_button("⬇️ 전체 현황 CSV",
                               data=df_csv.to_csv(index=False).encode("utf-8-sig"),
                               file_name=f"평가현황_{datetime.now().strftime('%Y%m%d')}.csv",
                               mime="text/csv")
            # 최종 집계표 엑셀
            summary_bytes = generate_summary_excel(evaluations, tasks_data, users, profiles, evaluatees)
            col_dl2.download_button("📊 최종 집계표 다운로드",
                               data=summary_bytes,
                               file_name=f"근무성적평정_결과집계표_{datetime.now().strftime('%Y%m%d')}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               type="primary")

    # ── 탭2 직무수행태도 ──────────────────────
    with tab2:
        st.subheader("🧘 직무수행태도 (C, 5점) 입력")
        st.caption("감점 항목 건수 입력 → 5점에서 자동 차감")
        st.divider()
        for uid, u in evaluatees.items():
            p    = profiles.get(uid, {})
            team = p.get("team", u.get("team",""))
            dept = get_team_dept(team)
            ev   = evaluations.get(uid, {})
            ded  = ev.get("deductions", {})
            att  = calc_attitude_score(ded)
            with st.expander(
                f"👤 {u.get('name','')}  ({dept}/{team}/{p.get('position',u.get('position',''))}) "
                f"— 태도점수: {att:.2f}점 / 5점", expanded=False
            ):
                with st.form(f"ded_{uid}"):
                    new_ded = {}
                    cl, cr = st.columns(2)
                    left_items  = list(DEDUCTION_RATES.items())[:6]
                    right_items = list(DEDUCTION_RATES.items())[6:]
                    with cl:
                        for item,(unit,rate) in left_items:
                            v = st.number_input(f"{item} ({unit}당 -{rate}점)",
                                                0, value=int(ded.get(item,0)), step=1,
                                                key=f"ded_{uid}_{item}")
                            new_ded[item] = v
                    with cr:
                        for item,(unit,rate) in right_items:
                            v = st.number_input(f"{item} ({unit}당 -{rate}점)",
                                                0, value=int(ded.get(item,0)), step=1,
                                                key=f"ded_{uid}_{item}")
                            new_ded[item] = v
                    preview = calc_attitude_score(new_ded)
                    st.markdown(f"총 감점: **{round(5-preview,2)}점** | 직무수행태도: **{preview:.2f}점**")
                    if st.form_submit_button("💾 저장", type="primary", use_container_width=True):
                        save_evaluation(uid, "deductions", {
                            **new_ded,
                            "deductions_updated": datetime.now().strftime("%Y-%m-%d %H:%M")
                        })
                        st.success(f"✅ {u.get('name','')} 직무수행태도 저장 완료."); st.rerun()

    # ── 탭3 직급별 순위 ───────────────────────
    with tab3:
        st.subheader("🏆 직급별 전사 통합 순위")
        st.caption("해당 직책 기준 모든 차수 평가 + 직무수행태도까지 완료된 피평가자만 집계됩니다.")
        df_rank = calc_grade_rankings(evaluations, tasks_data)

        if df_rank.empty:
            st.info("아직 전체 평가가 완료된 피평가자가 없습니다.")
        else:
            for grd in GRADES:
                df_g = df_rank[df_rank["직급"]==grd].copy()
                if df_g.empty: continue
                st.markdown(f"### 📌 {grd} 순위")
                # 존재하는 환산 컬럼만 표시
                base_cols = ["직급순위","이름","소속부","소속팀","직책","최종점수"]
                stage_cols = [c for c in df_g.columns if "환산" in c]
                disp = [c for c in base_cols + stage_cols if c in df_g.columns]
                st.dataframe(df_g[disp].rename(columns={"직급순위":"순위"}),
                             use_container_width=True, hide_index=True)
                st.divider()

            all_cols = [c for c in df_rank.columns if c != "uid"]
            st.download_button("⬇️ 직급별 순위 CSV",
                               data=df_rank[all_cols].to_csv(index=False).encode("utf-8-sig"),
                               file_name=f"직급별순위_{datetime.now().strftime('%Y%m%d')}.csv",
                               mime="text/csv")

    # ── 탭4 계정 관리 ─────────────────────────
    with tab4:
        st.subheader("🧑‍💼 계정·조직 관리")
        sub1, sub2, sub3 = st.tabs(["📥 엑셀 일괄 생성","✏️ 개별 계정 추가","📌 현재 계정 목록"])

        with sub1:
            col_dl, col_up = st.columns(2)
            with col_dl:
                with st.container(border=True):
                    st.markdown("**① 템플릿 다운로드**")
                    tpl = Path("account_template.xlsx")
                    if tpl.exists():
                        with open(tpl,"rb") as f:
                            st.download_button("⬇️ 템플릿 다운로드", data=f,
                                               file_name="계정생성_템플릿.xlsx",
                                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                               type="primary", use_container_width=True)
                    else:
                        st.warning("account_template.xlsx 를 앱 폴더에 넣어주세요.")
                    st.markdown("""
**컬럼 순서 (계정생성 시트)**
아이디 / 비밀번호 / 이름 / 직책 / 직급 / 소속팀 / 이메일
- 직책: 대표이사·본부장·부장·팀장·팀원
- 직급: 2급·3급·4급·5급·6급·공무직 (공무직은 직책 자동 팀원 지정)
- 소속팀: 행정지원팀·재정관리팀·생산관리팀·도담점·아름점·새롬점·소담점·직매장행정팀·공공급식팀
                    """)
            with col_up:
                with st.container(border=True):
                    st.markdown("**② 작성 완료 엑셀 업로드**")
                    xl = st.file_uploader("xlsx", type=["xlsx"],
                                          label_visibility="collapsed", key="acct_xl")

            if xl:
                fb = xl.read()
                ee_new, ev_new, errors = parse_account_excel_v3(fb)
                st.divider(); st.subheader("업로드 미리보기")
                if errors:
                    with st.expander(f"⚠️ 오류 {len(errors)}건", expanded=True):
                        for e in errors: st.warning(e)
                existing = get_users()
                for label, acc_dict in [("피평가자",ee_new),("평가자",ev_new)]:
                    if not acc_dict: continue
                    st.markdown(f"**{label} {len(acc_dict)}명**")
                    rows = [{"상태":"⚠️중복" if u in existing else "✅신규",
                             "아이디":u,"이름":v["name"],
                             "직책":v.get("position",""),"직급":v.get("grade",""),
                             "소속팀":v.get("team","")} for u,v in acc_dict.items()]
                    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

                if (ee_new or ev_new) and st.button("✅ 계정 생성 확정", type="primary"):
                    added_ee = added_ev = skipped = 0
                    for u, v in {**ee_new, **ev_new}.items():
                        if u not in existing:
                            save_user(u, v)
                            if v.get("role")=="evaluatee": added_ee+=1
                            else: added_ev+=1
                        else: skipped+=1
                    st.success(f"완료! 피평가자 {added_ee}명, 평가자 {added_ev}명 | 스킵 {skipped}명")
                    st.rerun()

        with sub2:
            st.markdown("**새 계정 추가**")
            with st.form("add_user_form"):
                c1, c2 = st.columns(2)
                new_uid   = c1.text_input("아이디 *",   key="nu_uid")
                new_pw    = c2.text_input("비밀번호 *", type="password", key="nu_pw")
                new_name  = c1.text_input("이름 *",     key="nu_name")
                new_role  = c2.selectbox("역할", ["evaluatee (피평가자)","evaluator (평가자)"], key="nu_role")
                new_pos   = c1.selectbox("직책 *",      POSITIONS, key="nu_pos")
                new_grd   = c2.selectbox("직급 *",      ["(선택)"] + GRADES, key="nu_grd")
                new_dept  = c1.selectbox("소속 부",     ["(해당없음)"] + ALL_DEPTS, key="nu_dept")
                # 선택된 부에 속한 팀만 표시
                dept_teams = ORG.get(new_dept, ALL_TEAMS) if new_dept != "(해당없음)" else ALL_TEAMS
                new_team  = c2.selectbox("소속 팀",     ["(해당없음)"] + dept_teams, key="nu_team")
                new_email = c1.text_input("이메일",     key="nu_email")

                # 직책별 안내
                guide = []
                if new_pos == "대표이사":  guide.append("직급·소속 부·팀 없음")
                elif new_pos == "본부장":  guide.append("소속 부·팀 없음")
                elif new_pos == "부장":    guide.append("소속 팀 없음")
                if new_grd == "공무직":    guide.append("공무직 → 직책 팀원 자동 지정")
                if guide: st.info(" | ".join(guide))

                submitted = st.form_submit_button("➕ 계정 추가", type="primary", use_container_width=True)

            if submitted:
                final_pos  = "팀원" if new_grd == "공무직" else new_pos
                final_grd  = "" if new_pos == "대표이사" else ("" if new_grd == "(선택)" else new_grd)
                final_dept = "" if new_pos in ("대표이사","본부장") else ("" if new_dept == "(해당없음)" else new_dept)
                final_team = "" if new_pos in ("대표이사","본부장","부장") else ("" if new_team == "(해당없음)" else new_team)

                if not new_uid or not new_pw or not new_name:
                    st.error("아이디·비밀번호·이름은 필수입니다.")
                elif not is_valid_id(new_uid):
                    st.error("아이디에 공백이나 특수문자는 사용할 수 없습니다.")
                elif new_pos != "대표이사" and not final_grd:
                    st.error("직급을 선택하세요.")
                elif new_pos in ("팀장","팀원") and not final_team:
                    st.error("팀장·팀원은 소속 팀이 필수입니다.")
                elif new_pos == "부장" and not final_dept:
                    st.error("부장은 소속 부가 필수입니다.")
                else:
                    users = get_users()
                    if new_uid in users:
                        st.error("이미 존재하는 아이디입니다.")
                    else:
                        users[new_uid] = {
                            "password": hash_pw(new_pw),
                            "role":     "evaluatee" if "evaluatee" in new_role else "evaluator",
                            "name":     new_name,   "position": final_pos,
                            "grade":    final_grd,  "team":     final_team,
                            "dept":     final_dept, "email":    new_email,
                        }
                        save_user(new_uid, users[new_uid])
                        st.success(f"✅ {new_name} 계정 추가 완료!"); st.rerun()

        with sub3:
            all_u = get_users(); all_p = get_profiles()

            # 삭제 확인 상태 관리
            if "delete_confirm" not in st.session_state:
                st.session_state.delete_confirm = None
            if "reset_confirm" not in st.session_state:
                st.session_state.reset_confirm = None

            # 삭제 확인 팝업
            if st.session_state.delete_confirm:
                del_uid  = st.session_state.delete_confirm
                del_name = all_u.get(del_uid, {}).get("name", del_uid)
                st.warning(f"⚠️ **'{del_name}' ({del_uid})** 계정을 삭제하시겠습니까?\n\n관련 평가 데이터·과제·인적사항도 함께 삭제됩니다.")
                cc1, cc2, _ = st.columns([1, 1, 3])
                if cc1.button("🗑️ 삭제 확정", type="primary", key="confirm_del"):
                    delete_user(del_uid)
                    delete_pdf(del_uid)
                    st.session_state.delete_confirm = None
                    st.success(f"✅ '{del_name}' 계정이 삭제되었습니다.")
                    st.rerun()
                if cc2.button("취소", key="cancel_del"):
                    st.session_state.delete_confirm = None
                    st.rerun()
                st.divider()

            # 초기화 확인 팝업
            if st.session_state.reset_confirm:
                rst_uid  = st.session_state.reset_confirm
                rst_name = all_u.get(rst_uid, {}).get("name", rst_uid)
                st.warning(
                    f"⚠️ **'{rst_name}' ({rst_uid})** 의 평가를 초기화하시겠습니까?\n\n"
                    f"인적사항·담당과제·평가점수·등급이 모두 삭제됩니다. 계정은 유지됩니다."
                )
                rc1, rc2, _ = st.columns([1, 1, 3])
                if rc1.button("🔄 초기화 확정", type="primary", key="confirm_reset"):
                    reset_evaluation(rst_uid)
                    st.session_state.reset_confirm = None
                    st.success(f"✅ '{rst_name}' 의 평가 데이터가 초기화되었습니다.")
                    st.rerun()
                if rc2.button("취소", key="cancel_reset"):
                    st.session_state.reset_confirm = None
                    st.rerun()
                st.divider()

            # 부별 계정 목록 + 삭제·초기화 버튼
            for dept, teams in ORG.items():
                st.markdown(f"#### 🏢 {dept}")
                dept_rows = []
                for uid, u in all_u.items():
                    if u.get("role") not in ("evaluatee", "evaluator"): continue
                    p    = all_p.get(uid, {})
                    team = p.get("team", u.get("team", ""))
                    if get_team_dept(team) != dept: continue
                    dept_rows.append((uid, u, p, team))

                if not dept_rows:
                    st.caption("등록된 계정 없음")
                    continue

                # 헤더
                hc = st.columns([1.5, 2, 1.5, 1.5, 1.5, 2, 1, 1])
                for h, label in zip(hc, ["아이디","이름","역할","직책","직급","소속팀","초기화","삭제"]):
                    h.markdown(f"**{label}**")
                st.divider()

                for uid, u, p, team in dept_rows:
                    role_str = "피평가자" if u["role"] == "evaluatee" else "평가자"
                    rc = st.columns([1.5, 2, 1.5, 1.5, 1.5, 2, 1, 1])
                    rc[0].caption(uid)
                    rc[1].markdown(u.get("name", ""))
                    rc[2].caption(role_str)
                    rc[3].caption(p.get("position", u.get("position", "")))
                    rc[4].caption(p.get("grade", u.get("grade", "")))
                    rc[5].caption(team)
                    if rc[6].button("🔄", key=f"rst_{uid}",
                                    help=f"{u.get('name','')} 평가 초기화"):
                        st.session_state.reset_confirm = uid
                        st.rerun()
                    if rc[7].button("🗑️", key=f"del_{uid}",
                                    help=f"{u.get('name','')} 계정 삭제"):
                        st.session_state.delete_confirm = uid
                        st.rerun()

            # ── 소속 없는 계정 (대표이사·본부장) & 기타 ──
            st.markdown("#### 👑 임원 (대표이사·본부장) 및 기타")
            etc_rows = []
            for uid, u in all_u.items():
                if u.get("role") not in ("evaluatee", "evaluator"): continue
                p    = all_p.get(uid, {})
                team = p.get("team", u.get("team", ""))
                pos  = p.get("position", u.get("position", ""))
                if pos in ("대표이사", "본부장") or get_team_dept(team) == "":
                    etc_rows.append((uid, u, p, team))

            if etc_rows:
                hc = st.columns([1.5, 2, 1.5, 1.5, 1.5, 2, 1, 1])
                for h, label in zip(hc, ["아이디","이름","역할","직책","직급","소속팀","초기화","삭제"]):
                    h.markdown(f"**{label}**")
                st.divider()
                for uid, u, p, team in etc_rows:
                    role_str = "피평가자" if u["role"] == "evaluatee" else "평가자"
                    rc = st.columns([1.5, 2, 1.5, 1.5, 1.5, 2, 1, 1])
                    rc[0].caption(uid)
                    rc[1].markdown(u.get("name", ""))
                    rc[2].caption(role_str)
                    rc[3].caption(p.get("position", u.get("position", "")))
                    rc[4].caption(p.get("grade", u.get("grade", "")))
                    rc[5].caption(team if team else "-")
                    if rc[6].button("🔄", key=f"rst_{uid}",
                                    help=f"{u.get('name','')} 평가 초기화"):
                        st.session_state.reset_confirm = uid
                        st.rerun()
                    if rc[7].button("🗑️", key=f"del_{uid}",
                                    help=f"{u.get('name','')} 계정 삭제"):
                        st.session_state.delete_confirm = uid
                        st.rerun()
            else:
                st.caption("해당 계정 없음")



# ══════════════════════════════════════════════════════════════
# 최종 집계표 엑셀 생성
# ══════════════════════════════════════════════════════════════
def generate_summary_excel(evaluations, tasks_data, users, profiles, evaluatees) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    GRADE_ORDER = ["2급","3급","4급","5급","6급","공무직"]
    GRADE_BG    = {"2급":"E8F4FD","3급":"E8F4FD","4급":"E8F4FD",
                   "5급":"FFF8E8","6급":"FFF0E8","공무직":"F0E8FF"}
    GRADE_HDR   = {"2급":"1F4E79","3급":"1F4E79","4급":"1F4E79",
                   "5급":"7B6000","6급":"833C00","공무직":"4B0082"}

    def thin():
        s = Side(style="thin", color="AAAAAA")
        return Border(left=s, right=s, top=s, bottom=s)

    def hc(ws, r, c, val, bg="1F4E79", fg="FFFFFF", bold=True, sz=9):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = Font(name="맑은 고딕", bold=bold, color=fg, size=sz)
        cell.fill = PatternFill("solid", start_color=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin()
        return cell

    def dc(ws, r, c, val, bg="FFFFFF", bold=False, sz=9, align_h="center"):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = Font(name="맑은 고딕", bold=bold, size=sz)
        cell.fill = PatternFill("solid", start_color=bg)
        cell.alignment = Alignment(horizontal=align_h, vertical="center", wrap_text=True)
        cell.border = thin()
        return cell

    def mc(ws, r1, c1, r2, c2):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

    def fmt(v):
        if isinstance(v, (int, float)):
            return round(v, 2)
        return "-"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "최종집계"
    ws.sheet_view.showGridLines = False

    # 열 너비
    widths = [4, 9, 7, 11, 10, 6, 8, 7, 7, 7, 7, 7, 7, 7, 7, 7, 7, 7, 5]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 타이틀
    mc(ws,1,1,1,19)
    hc(ws,1,1,f"■ {datetime.now().year}년 근무성적평정 결과집계표",
       bg="FFFFFF", fg="1F4E79", bold=True, sz=13)
    ws.row_dimensions[1].height = 26

    # 헤더 3행 구조
    for r in [2,3,4]: ws.row_dimensions[r].height = 22

    # 고정 컬럼 (2~4행 병합)
    for c, label in enumerate(["순번","성명","직급","소속부","소속팀","직책","비고"],1):
        mc(ws,2,c,4,c); hc(ws,2,c,label,bg="1F4E79")

    # 평가점수 그룹
    mc(ws,2,8,2,11); hc(ws,2,8,"평가점수 (95점 만점)",bg="2E75B6")
    for c,label in enumerate(["1차\n평가","2차\n평가","3차\n평가","4차\n평가"],8):
        mc(ws,3,c,4,c); hc(ws,3,c,label,bg="2E75B6")

    # 환산점수 그룹
    mc(ws,2,12,2,15); hc(ws,2,12,"환산점수",bg="375623")
    for c,label in enumerate(["1차\n환산","2차\n환산","3차\n환산","4차\n환산"],12):
        mc(ws,3,c,4,c); hc(ws,3,c,label,bg="375623")

    # 소계·태도·합계·순위
    mc(ws,2,16,4,16); hc(ws,2,16,"소계\n(95점)",bg="C00000")
    mc(ws,2,17,4,17); hc(ws,2,17,"직무수행\n태도(5점)",bg="C00000")
    mc(ws,2,18,4,18); hc(ws,2,18,"계\n(100점)",bg="C00000")
    mc(ws,2,19,4,19); hc(ws,2,19,"순위",bg="1F4E79")

    # 데이터 수집 및 정렬
    all_rows = []
    for uid, u in evaluatees.items():
        p       = profiles.get(uid, {})
        ev      = evaluations.get(uid, {})
        ee_pos  = p.get("position", u.get("position","팀원"))
        ee_grd  = p.get("grade", u.get("grade",""))
        ee_stages = get_stage_order_for_ee(ee_pos)
        ee_weights= get_eval_weights_for_ee(ee_pos)
        tasks   = tasks_data.get(uid, {}).get("tasks", [])

        # 차수별 점수 (A+B 소계 기준 95점 환산)
        stage_scores = {}
        for st_ in ["1차","2차","3차","4차"]:
            ev_st = ev.get(st_, {})
            if ev_st:
                a = calc_task_score(tasks, ev_st)
                b = calc_ability_score(ev_st)
                stage_scores[st_] = round(a + b, 2)  # 95점 만점

        # 환산점수 (비중 적용)
        stage_conv = {}
        for st_ in ee_stages:
            sc = stage_scores.get(st_)
            if sc is not None:
                stage_conv[st_] = round(sc * ee_weights[st_], 2)

        subtotal  = round(sum(stage_conv.values()), 2) if stage_conv else None
        deductions= ev.get("deductions", {})
        attitude  = calc_attitude_score(deductions) if ev.get("deductions") is not None else None
        total     = round(subtotal + (attitude or 0), 2) if subtotal is not None else None

        all_rows.append({
            "uid":      uid,
            "name":     u.get("name",""),
            "grade":    ee_grd,
            "dept":     get_team_dept(p.get("team", u.get("team",""))),
            "team":     p.get("team", u.get("team","")),
            "position": ee_pos,
            "s1": stage_scores.get("1차"), "s2": stage_scores.get("2차"),
            "s3": stage_scores.get("3차"), "s4": stage_scores.get("4차"),
            "e1": stage_conv.get("1차"),  "e2": stage_conv.get("2차"),
            "e3": stage_conv.get("3차"),  "e4": stage_conv.get("4차"),
            "subtotal": subtotal,
            "attitude": attitude,
            "total":    total,
        })

    # 직급별 순위 계산
    from collections import defaultdict
    by_grade = defaultdict(list)
    for r in all_rows:
        by_grade[r["grade"]].append(r)

    for grd, grp in by_grade.items():
        completed = [r for r in grp if r["total"] is not None]
        completed_sorted = sorted(completed, key=lambda x: x["total"], reverse=True)
        rank = 1
        prev_score = None
        for i, r in enumerate(completed_sorted):
            if r["total"] != prev_score:
                rank = i + 1
            r["rank"] = rank
            prev_score = r["total"]
        for r in grp:
            if r.get("rank") is None:
                r["rank"] = "-"

    # 정렬: 상위직급부터 → 직급별 순위
    def sort_key(r):
        g = r.get("grade","공무직")
        idx = GRADE_ORDER.index(g) if g in GRADE_ORDER else 99
        rank = r.get("rank", 99)
        return (idx, rank if isinstance(rank, int) else 99)

    all_rows.sort(key=sort_key)

    # 데이터 작성
    data_row = 5
    prev_grade = None
    grade_no = {}

    for r in all_rows:
        g   = r.get("grade","공무직")
        bg  = GRADE_BG.get(g, "FFFFFF")
        hbg = GRADE_HDR.get(g, "333333")

        if g != prev_grade:
            mc(ws, data_row, 1, data_row, 19)
            hc(ws, data_row, 1, f"▶  {g}", bg=hbg, sz=9)
            ws.row_dimensions[data_row].height = 16
            data_row += 1
            prev_grade = g
            grade_no[g] = 1

        no = grade_no.get(g, 1); grade_no[g] = no + 1

        dc(ws, data_row, 1,  no,                    bg=bg)
        dc(ws, data_row, 2,  r.get("name",""),       bg=bg, bold=True)
        dc(ws, data_row, 3,  g,                      bg=bg)
        dc(ws, data_row, 4,  r.get("dept",""),       bg=bg, align_h="left")
        dc(ws, data_row, 5,  r.get("team",""),       bg=bg, align_h="left")
        dc(ws, data_row, 6,  r.get("position",""),   bg=bg)
        dc(ws, data_row, 7,  "",                     bg="FFFDE7")  # 비고(수기)

        for c, key in enumerate(["s1","s2","s3","s4"], 8):
            v = r.get(key)
            dc(ws, data_row, c, fmt(v),
               bg="EBF3FB" if v is not None else "F5F5F5")

        for c, key in enumerate(["e1","e2","e3","e4"], 12):
            v = r.get(key)
            dc(ws, data_row, c, fmt(v),
               bg="E8F5E9" if v is not None else "F5F5F5")

        dc(ws, data_row, 16, fmt(r.get("subtotal")), bg="FFF2CC", bold=True)
        dc(ws, data_row, 17, fmt(r.get("attitude")), bg="FFF2CC")
        dc(ws, data_row, 18, fmt(r.get("total")),    bg="FFD700", bold=True)
        dc(ws, data_row, 19, r.get("rank","-"),      bg="E8D5F5", bold=True)

        ws.row_dimensions[data_row].height = 18
        data_row += 1

    # 합계 행
    mc(ws, data_row, 1, data_row, 7)
    hc(ws, data_row, 1, f"총  {len(all_rows)}명", bg="D6E4F0", fg="1F4E79", sz=10)
    for c in range(8, 20):
        dc(ws, data_row, c, "", bg="D6E4F0")
    ws.row_dimensions[data_row].height = 20

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ══════════════════════════════════════════════════════════════
# 엑셀 계정 파싱 v3
# ══════════════════════════════════════════════════════════════
def parse_account_excel_v3(file_bytes):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ee_new, ev_new, errors = {}, {}, []
    if "계정생성" not in wb.sheetnames:
        errors.append("'계정생성' 시트를 찾을 수 없습니다."); return ee_new, ev_new, errors

    ws = wb["계정생성"]; cur_role = None
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        first = str(row[0]).strip() if row[0] else ""
        if "피평가자" in first and "evaluatee" in first.lower(): cur_role="evaluatee"; continue
        if "평가자"  in first and "evaluator" in first.lower():  cur_role="evaluator"; continue
        if "아이디"  in first: continue

        def cell(i): return str(row[i]).strip() if len(row)>i and row[i] not in (None,"None") else ""
        uid_raw=cell(0); pw_raw=cell(1); name_raw=cell(2)
        position=cell(3); grade=cell(4); team=cell(5); email=cell(6)

        if not uid_raw: continue
        if not is_valid_id(uid_raw): errors.append(f"행{r_idx}: '{uid_raw}' — 아이디에 공백이나 특수문자가 포함되어 있습니다."); continue
        if not pw_raw:   errors.append(f"행{r_idx}: '{uid_raw}' 비밀번호 없음"); continue
        if not name_raw: errors.append(f"행{r_idx}: '{uid_raw}' 이름 없음"); continue
        if grade    not in GRADES:    errors.append(f"행{r_idx}: '{uid_raw}' 직급 오류({grade})"); continue
        # 공무직은 직책 부여 불가 → 팀원 자동 지정
        if grade == "공무직":
            position = "팀원"
        elif position not in POSITIONS:
            errors.append(f"행{r_idx}: '{uid_raw}' 직책 오류({position})"); continue
        if team not in ALL_TEAMS: errors.append(f"행{r_idx}: '{uid_raw}' 팀 오류({team})"); continue

        entry = {"password":hash_pw(pw_raw),"name":name_raw,"position":position,
                 "grade":grade,"team":team,"dept":get_team_dept(team),"email":email}
        if cur_role=="evaluatee": entry["role"]="evaluatee"; ee_new[uid_raw]=entry
        elif cur_role=="evaluator": entry["role"]="evaluator"; ev_new[uid_raw]=entry

    return ee_new, ev_new, errors


# ══════════════════════════════════════════════════════════════
# 개인 평정표 엑셀 생성
# ══════════════════════════════════════════════════════════════
def generate_eval_excel(ee_id, ee_info, evaluations, tasks_data):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    profiles = get_profiles()
    p      = profiles.get(ee_id, {})
    tasks  = tasks_data.get(ee_id, {}).get("tasks", [])
    ev     = evaluations.get(ee_id, {})
    result = calc_final(ee_id, evaluations, tasks_data)
    종합   = result.get("종합", {})

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title="근무성적평정표"
    ws.sheet_view.showGridLines = False

    thin = Side(style="thin", color="AAAAAA")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hc(r,c,val,bg="1F4E79",fg="FFFFFF",bold=True,sz=10):
        cell=ws.cell(row=r,column=c,value=val)
        cell.font=Font(name="맑은 고딕",bold=bold,color=fg,size=sz)
        cell.fill=PatternFill("solid",start_color=bg)
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        cell.border=bdr; return cell

    def dc(r,c,val,bold=False,align="left",bg="FFFFFF",sz=10):
        cell=ws.cell(row=r,column=c,value=val)
        cell.font=Font(name="맑은 고딕",bold=bold,size=sz)
        cell.fill=PatternFill("solid",start_color=bg)
        cell.alignment=Alignment(horizontal=align,vertical="center",wrap_text=True)
        cell.border=bdr; return cell

    def mc(r1,c1,r2,c2): ws.merge_cells(start_row=r1,start_column=c1,end_row=r2,end_column=c2)

    for i,w in enumerate([4,6,14,10,10,18,8,8,8,8],1):
        ws.column_dimensions[get_column_letter(i)].width=w

    team = p.get("team", ee_info.get("team",""))
    dept = get_team_dept(team)

    # 제목
    mc(1,1,1,10); hc(1,1,"근무성적 평정표",sz=14); ws.row_dimensions[1].height=35

    # 인적사항
    mc(2,1,2,10); hc(2,1,"□ 인적 사항",bg="2E75B6",sz=11)
    hc(3,1,"성명",bg="D6E4F0",fg="000000"); mc(3,2,3,3); dc(3,2,ee_info.get("name",""))
    hc(3,4,"소속부",bg="D6E4F0",fg="000000"); mc(3,5,3,6); dc(3,5,dept)
    hc(3,7,"소속팀",bg="D6E4F0",fg="000000"); mc(3,8,3,10); dc(3,8,team)
    hc(4,1,"직책",bg="D6E4F0",fg="000000"); mc(4,2,4,3)
    dc(4,2,p.get("position",ee_info.get("position","")))
    hc(4,4,"직급",bg="D6E4F0",fg="000000"); mc(4,5,4,6)
    dc(4,5,p.get("grade",ee_info.get("grade","")))
    hc(4,7,"입사일",bg="D6E4F0",fg="000000"); mc(4,8,4,10); dc(4,8,p.get("입사일",""))
    mc(5,1,5,10); dc(5,1,f"담당업무: {p.get('담당업무','')}",bg="EBF3FB",sz=9)

    row=6
    ee_pos    = p.get("position", ee_info.get("position","팀원"))
    ee_stages = get_stage_order_for_ee(ee_pos)
    ee_weights= get_eval_weights_for_ee(ee_pos)
    users_all = get_users()
    stage_labels = {"1차":"팀장","2차":"부장","3차":"본부장","4차":"대표이사"}

    # ── 2. 근무실적 평정 (A, 60점) ──────────────────────────────
    mc(row,1,row,10); hc(row,1,"2. 근무실적 평정 (A, 60점)",bg="2E75B6",sz=11)
    ws.row_dimensions[row].height=22; row+=1
    # 헤더
    mc(row,1,row,2); hc(row,1,"구분",bg="375623",sz=9)
    hc(row,3,"과제명",bg="375623",sz=9); hc(row,4,"비중",bg="375623",sz=9)
    hc(row,5,"주요실적",bg="375623",sz=9)
    col_idx=6
    for st_ in ee_stages:
        wt=ee_weights[st_]
        mc(row,col_idx,row,col_idx+1)
        hc(row,col_idx,f"{st_}({int(wt*100)}%)\n평가/환산",bg="1F4E79",sz=8)
        col_idx+=2
    ws.row_dimensions[row].height=28; row+=1
    # 과제별 행
    for t in tasks:
        mc(row,1,row,2); dc(row,1,t["type"],align="center")
        dc(row,3,t["title"]); dc(row,4,f"{t['weight']:.0%}",align="center")
        dc(row,5,t.get("result",""),bg="FAFAFA")
        col_idx=6
        for st_ in ee_stages:
            ev_st=ev.get(st_,{}); sc=ev_st.get("tasks",{}).get(t["id"],"")
            conv=round(sc*t["weight"]*6,2) if isinstance(sc,(int,float)) else "-"
            dc(row,col_idx,sc if sc!="" else "-",align="center",bg="FFF2CC" if isinstance(sc,(int,float)) else "FFE8E8")
            dc(row,col_idx+1,conv,align="center",bg="FFF2CC"); col_idx+=2
        ws.row_dimensions[row].height=38; row+=1
    # 합계
    mc(row,1,row,4); hc(row,1,"합 계 (60점)",bg="D6E4F0",fg="000000"); dc(row,5,"")
    col_idx=6
    for st_ in ee_stages:
        a=calc_task_score(tasks,ev.get(st_,{}))
        dc(row,col_idx,"",bg="F0F0F0"); dc(row,col_idx+1,f"{a:.1f}",align="center",bold=True,bg="FFF2CC"); col_idx+=2
    ws.row_dimensions[row].height=18; row+=1

    # ── 3-1. 직무수행능력 (B, 35점) ─────────────────────────────
    mc(row,1,row,10); hc(row,1,"3-1. 직무수행능력 (B, 35점)",bg="2E75B6",sz=11)
    ws.row_dimensions[row].height=22; row+=1
    mc(row,1,row,2); hc(row,1,"연번",bg="375623",sz=9)
    hc(row,3,"평정요소",bg="375623",sz=9); hc(row,4,"배점",bg="375623",sz=9); hc(row,5,"정의",bg="375623",sz=9)
    col_idx=6
    for st_ in ee_stages:
        mc(row,col_idx,row,col_idx+1); hc(row,col_idx,st_,bg="1F4E79",sz=8); col_idx+=2
    ws.row_dimensions[row].height=18; row+=1
    for i,(ab_name,ab_max,ab_def) in enumerate(ABILITY_ITEMS,1):
        mc(row,1,row,2); dc(row,1,i,align="center")
        dc(row,3,ab_name); dc(row,4,ab_max,align="center"); dc(row,5,ab_def.replace("\n"," "),bg="FAFAFA")
        col_idx=6
        for st_ in ee_stages:
            sc=ev.get(st_,{}).get("ability",{}).get(ab_name,"")
            mc(row,col_idx,row,col_idx+1)
            dc(row,col_idx,sc if sc!="" else "-",align="center",bg="FFF2CC" if isinstance(sc,(int,float)) else "FFE8E8")
            col_idx+=2
        ws.row_dimensions[row].height=26; row+=1
    mc(row,1,row,4); hc(row,1,"합 계 (35점)",bg="D6E4F0",fg="000000"); dc(row,5,"")
    col_idx=6
    for st_ in ee_stages:
        b=calc_ability_score(ev.get(st_,{}))
        mc(row,col_idx,row,col_idx+1); dc(row,col_idx,f"{b:.0f}",align="center",bold=True,bg="FFF2CC"); col_idx+=2
    ws.row_dimensions[row].height=18; row+=1

    # ── 3-2. 직무수행태도 (C, 5점) ──────────────────────────────
    mc(row,1,row,10); hc(row,1,"3-2. 직무수행태도 (C, 5점) — 총괄 관리자 입력",bg="2E75B6",sz=11)
    ws.row_dimensions[row].height=22; row+=1
    deductions=ev.get("deductions",{}); total_ded=0.0
    items_l=list(DEDUCTION_RATES.items()); left_i=items_l[:6]; right_i=items_l[6:]
    mc(row,1,row,3); hc(row,1,"감점 항목",bg="375623",sz=9)
    hc(row,4,"기준",bg="375623",sz=9); hc(row,5,"건수",bg="375623",sz=9); hc(row,6,"감점",bg="C00000",sz=9)
    mc(row,7,row,8); hc(row,7,"감점 항목",bg="375623",sz=9)
    hc(row,9,"기준",bg="375623",sz=9); hc(row,10,"건수",bg="375623",sz=9)
    ws.row_dimensions[row].height=16; row+=1
    for i in range(max(len(left_i),len(right_i))):
        if i<len(left_i):
            item,(unit,rate)=left_i[i]; cnt=deductions.get(item,0); ded=round(cnt*rate,2); total_ded+=ded
            mc(row,1,row,3); dc(row,1,item); dc(row,4,f"{unit}당 -{rate}점",align="center")
            dc(row,5,cnt,align="center"); dc(row,6,f"-{ded}점" if cnt>0 else "-",align="center",bg="FFE8E8" if cnt>0 else "FFFFFF")
        if i<len(right_i):
            item2,(unit2,rate2)=right_i[i]; cnt2=deductions.get(item2,0)
            mc(row,7,row,8); dc(row,7,item2); dc(row,9,f"{unit2}당 -{rate2}점",align="center"); dc(row,10,cnt2,align="center")
        ws.row_dimensions[row].height=16; row+=1
    att_c=calc_attitude_score(deductions)
    mc(row,1,row,5); hc(row,1,f"직무수행태도 합계  (총 감점: -{round(total_ded,2)}점)",bg="D6E4F0",fg="000000",sz=10)
    mc(row,6,row,10); dc(row,6,f"{att_c:.2f}점",align="center",bold=True,bg="FFF2CC",sz=11)
    ws.row_dimensions[row].height=20; row+=1

    # ── 4. 종합평가 ──────────────────────────────────────────────
    mc(row,1,row,10); hc(row,1,"4. 종합평가",bg="2E75B6",sz=11)
    ws.row_dimensions[row].height=22; row+=1
    mc(row,1,row,4); hc(row,1,"구분",bg="375623",sz=9); hc(row,5,"가중환산",bg="C00000",sz=9)
    col_idx=6
    for st_ in ee_stages:
        wt=ee_weights[st_]; mc(row,col_idx,row,col_idx+1); hc(row,col_idx,f"{st_}({int(wt*100)}%)",bg="1F4E79",sz=8); col_idx+=2
    ws.row_dimensions[row].height=18; row+=1
    result_data=calc_final(ee_id,evaluations,tasks_data); 종합=result_data.get("종합",{})
    for label,key_a,key_b in [("근무실적 평정(A, 60점)","A_가중","A"),("직무수행능력(B, 35점)","B_가중","B")]:
        mc(row,1,row,4); dc(row,1,label)
        dc(row,5,종합.get(key_a,0),align="center",bold=True,bg="FFF2CC")
        col_idx=6
        for st_ in ee_stages:
            val=result_data.get(st_,{}).get(key_b[-1],"-")
            mc(row,col_idx,row,col_idx+1); dc(row,col_idx,val,align="center"); col_idx+=2
        ws.row_dimensions[row].height=18; row+=1
    mc(row,1,row,4); hc(row,1,"소계 (A+B)",bg="D6E4F0",fg="000000",sz=10)
    dc(row,5,종합.get("AB_가중",0),align="center",bold=True,bg="FFF2CC")
    col_idx=6
    for st_ in ee_stages:
        ab=result_data.get(st_,{}).get("AB","-")
        mc(row,col_idx,row,col_idx+1); dc(row,col_idx,ab,align="center"); col_idx+=2
    ws.row_dimensions[row].height=18; row+=1
    mc(row,1,row,4); dc(row,1,"직무수행태도 (C, 5점)")
    dc(row,5,att_c,align="center",bold=True,bg="FFF2CC"); mc(row,6,row,10); dc(row,6,"")
    ws.row_dimensions[row].height=18; row+=1
    mc(row,1,row,4); hc(row,1,"최종합계 (A+B+C, 100점)",bg="1F4E79",sz=11)
    final_score=round(종합.get("AB_가중",0)+att_c,2)
    dc(row,5,final_score,align="center",bold=True,bg="FFD700",sz=13); mc(row,6,row,10); dc(row,6,"",bg="F0F0F0")
    ws.row_dimensions[row].height=26; row+=1
    assigned_grade=ev.get("assigned_grade","미확정")
    mc(row,1,row,4); hc(row,1,"확정 등급 (총괄 관리자 지정)",bg="1F4E79",sz=11)
    dc(row,5,assigned_grade,align="center",bold=True,bg="E8D5F5" if assigned_grade!="미확정" else "F5F5F5",sz=13)
    mc(row,6,row,10); dc(row,6,"",bg="F0F0F0"); ws.row_dimensions[row].height=26; row+=1

    # ── 5. 평정 의견 ─────────────────────────────────────────────
    mc(row,1,row,10); hc(row,1,"5. 평정 의견",bg="2E75B6",sz=11)
    ws.row_dimensions[row].height=22; row+=1
    for st_ in ee_stages:
        op=ev.get(st_,{}).get("opinion",""); ev_uid_=ev.get(st_,{}).get("evaluator_id","")
        ev_nm=users_all.get(ev_uid_,{}).get("name",ev_uid_)
        mc(row,1,row,2); hc(row,1,f"{st_}\n({stage_labels.get(st_,'')} {ev_nm})",bg="375623",sz=8)
        mc(row,3,row,10); dc(row,3,op,bg="FAFAFA"); ws.row_dimensions[row].height=26; row+=1

    # ── 시트2: 근무성적평정서(참고자료) ─────────────────────────
    sr = get_selfreport(ee_id)
    ws2 = wb.create_sheet("참고자료(근무성적평정서)")
    ws2.sheet_view.showGridLines = False

    for i, w in enumerate([4, 60], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    def hc2(r, val, bg="2E75B6", sz=11):
        ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        c = ws2.cell(row=r, column=1, value=val)
        c.font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=sz)
        c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = bdr
        ws2.row_dimensions[r].height = 24

    def dc2(r, label, value, bg_label="D6E4F0", bg_val="FFFFFF"):
        c1 = ws2.cell(row=r, column=1, value=label)
        c1.font = Font(name="맑은 고딕", bold=True, size=9)
        c1.fill = PatternFill("solid", start_color=bg_label)
        c1.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c1.border = bdr
        c2 = ws2.cell(row=r, column=2, value=value)
        c2.font = Font(name="맑은 고딕", size=10)
        c2.fill = PatternFill("solid", start_color=bg_val)
        c2.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c2.border = bdr

    # 타이틀
    ws2.merge_cells("A1:B1")
    t2 = ws2["A1"]
    t2.value = f"근무성적평정서 — {ee_info.get('name','')} ({p.get('grade','')} / {p.get('position','')})"
    t2.font = Font(name="맑은 고딕", bold=True, size=14, color="1F4E79")
    t2.fill = PatternFill("solid", start_color="D6E4F0")
    t2.alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[1].height = 32

    sr_row = 2

    # 자기계발 사항
    hc2(sr_row, "□ 자기계발 사항"); sr_row += 1
    dc2(sr_row, "1. 올해 교육연수를\n받거나 연구한 사항", sr.get("dev1","(미작성)"))
    ws2.row_dimensions[sr_row].height = max(80, len(sr.get("dev1","")) // 2 + 40)
    sr_row += 1
    dc2(sr_row, "2. 차기년도 희망\n교육연수 사항", sr.get("dev2","(미작성)"))
    ws2.row_dimensions[sr_row].height = max(80, len(sr.get("dev2","")) // 2 + 40)
    sr_row += 1

    # 다음연도 업무목표
    hc2(sr_row, "□ 다음연도 고과평가 대상 기간 중 추진하고자 하는 업무목표 (5개 이내)")
    sr_row += 1
    goals = sr.get("goals", [])
    for i, g in enumerate(goals, 1):
        if g:
            dc2(sr_row, f"목표 {i}", g)
            ws2.row_dimensions[sr_row].height = max(30, len(g) // 3 + 20)
            sr_row += 1
    if not any(goals):
        dc2(sr_row, "업무목표", "(미작성)")
        ws2.row_dimensions[sr_row].height = 30
        sr_row += 1

    # 희망부서 및 건의사항
    hc2(sr_row, "□ 희망부서 및 건의사항"); sr_row += 1
    dc2(sr_row, "희망부서 및\n건의사항", sr.get("suggestion","(미작성)"))
    ws2.row_dimensions[sr_row].height = max(80, len(sr.get("suggestion","")) // 2 + 40)
    sr_row += 1

    # 작성일
    ws2.merge_cells(f"A{sr_row}:B{sr_row}")
    c = ws2.cell(row=sr_row, column=1,
                 value=f"작성일: {sr.get('updated_at','미작성')}")
    c.font = Font(name="맑은 고딕", size=9, color="888888", italic=True)
    c.alignment = Alignment(horizontal="right", vertical="center")
    ws2.row_dimensions[sr_row].height = 20

    out=io.BytesIO(); wb.save(out); return out.getvalue()



# ══════════════════════════════════════════════════════════════
# 메인
# ══════════════════════════════════════════════════════════════
def main():
    st.set_page_config(page_title="근무성적 평정 시스템", page_icon="📋",
                       layout="wide", initial_sidebar_state="expanded")

    if not st.session_state.logged_in:
        show_login(); return

    users = get_users()
    u     = users.get(st.session_state.username, {})
    role  = st.session_state.role

    with st.sidebar:
        role_icon = {"evaluatee":"📄 피평가자","evaluator":"✍️ 평가자","admin":"🗂️ 총괄 관리자"}.get(role,"")
        st.markdown(f"### {role_icon}")
        st.markdown(f"**{st.session_state.name}** 님")
        if role == "evaluator":
            stage = get_stage_for_evaluator(st.session_state.username)
            st.caption(f"직책: {u.get('position','')} | 담당 차수: {stage}")
        if role in ("evaluatee","evaluator"):
            p    = get_profiles().get(st.session_state.username, {})
            team = p.get("team", u.get("team",""))
            st.caption(f"소속: {get_team_dept(team)} / {team}")
            st.caption(f"직급: {p.get('grade', u.get('grade',''))}")
        st.divider()

        # 비밀번호 변경 (관리자 제외)
        if role in ("evaluatee", "evaluator"):
            with st.expander("🔑 비밀번호 변경"):
                cur_pw  = st.text_input("현재 비밀번호", type="password", key="cur_pw")
                new_pw1 = st.text_input("새 비밀번호",   type="password", key="new_pw1")
                new_pw2 = st.text_input("새 비밀번호 확인", type="password", key="new_pw2")
                if st.button("변경", type="primary", use_container_width=True, key="pw_change"):
                    users_w = get_users()
                    uid_w   = st.session_state.username
                    if not cur_pw or not new_pw1 or not new_pw2:
                        st.error("모든 항목을 입력하세요.")
                    elif users_w.get(uid_w, {}).get("password") != hash_pw(cur_pw):
                        st.error("현재 비밀번호가 올바르지 않습니다.")
                    elif new_pw1 != new_pw2:
                        st.error("새 비밀번호가 일치하지 않습니다.")
                    elif len(new_pw1) < 4:
                        st.error("비밀번호는 4자 이상이어야 합니다.")
                    else:
                        users_w[uid_w]["password"] = hash_pw(new_pw1)
                        save_user(uid_w, users_w[uid_w])
                        st.success("✅ 비밀번호가 변경되었습니다.")

        st.divider()
        if st.button("🚪 로그아웃", use_container_width=True):
            for k in list(st.session_state.keys()): del st.session_state[k]
            st.rerun()

    if role=="evaluatee":   show_evaluatee()
    elif role=="evaluator": show_evaluator()
    elif role=="admin":     show_admin()
    else: st.error("알 수 없는 계정 유형입니다.")

if __name__ == "__main__":
    main()
